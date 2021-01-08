from flask import Flask, render_template, request, redirect, url_for , jsonify
from flaskext.mysql import MySQL
import xlrd
from openpyxl import load_workbook, Workbook
from werkzeug.utils import secure_filename
from sklearn.externals import joblib
import os
import pandas as pd
import json
import re
import nltk
import warnings
import numpy as np

import seaborn as sns
from argparse import Namespace
import matplotlib.pyplot as plt
from nltk.corpus import stopwords
from sklearn.utils import shuffle
from sklearn.feature_extraction.text import TfidfVectorizer

import re
import pdftotext
import pandas as pd
from pathlib import Path



model_name = "KNN-f1-0.851"
models= joblib.load("/home/ibu/Desktop/Annotated /test/models/upsamp/{}.{}".format(model_name,"pkl"))


app = Flask(__name__)


app.config['MYSQL_DATABASE_HOST'] = 'localhost'
app.config['MYSQL_DATABASE_USER'] = 'root'
app.config['MYSQL_DATABASE_PASSWORD'] = ''
app.config['MYSQL_DATABASE_DB'] = 'ADR_DEV'

mysql = MySQL(app)

@app.route('/')
def index():
    return render_template('blank.html')


@app.route('/upload', methods=['GET', 'POST'])
def up1():
    df = pd.DataFrame(columns = ['Progress Note', 'Diagnosis', 'Lab Order', 'Medication', 'Label'])

    file = request.files['inputfile']
    file_name = 'temp.'+file.filename.split(".")[1]
    file.save(secure_filename(file_name))
    print(file_name)
    with open(file_name, 'rb') as f:
        pdf = pdftotext.PDF(f)

    text = '\n\n'.join(pdf)
    with open(f'./output/file_name.text', 'w') as f:
        f.write(text)

    lines = []
    lab = []
    progress = []
    medication = []
    diagnosis = []
    with open(f'./output/file_name.text') as f:
        for line in f.readlines():
            lines.append(str(line).strip())
    lines = [sub.replace('*** End ***', '') for sub in lines]
    lines = list(filter(lambda x: not re.match('[0-9]{2}[\-,:][0-9]{2}[\-,:][0-9]{2}', x), lines))
    for (i, line) in enumerate(lines):
        sentence = str(line).strip()
        if (sentence.startswith('UHID')):
            lines[i] = ''
    titles = ['PROGRESS NOTES Doctor','LAB ORDER Doctor','MEDICATION ORDER Doctor','DIAGNOSIS Doctor']
    for (i, line) in enumerate(lines):
        sentence = str(line).strip()
        if any(sentence.startswith(x) for x in titles):
            notes = []
            if (sentence.startswith(titles[0])):
                notes = progress
            if (sentence.startswith(titles[1])):
                notes = lab
            if (sentence.startswith(titles[2])):
                notes = medication
            if (sentence.startswith(titles[3])):
                notes = diagnosis

            valid= True
            line_index = 1
            while valid:
                try:
                    line_state = any(lines[i+line_index].startswith(a) for a in titles)
                    if line_state:
                        valid = False
                    notes.append(lines[i+ line_index])
                    line_index += 1
                except:
                    break
    lab = ''.join([' '.join(lab)])
    progress = ''.join([' '.join(progress)])
    diagnosis = ''.join([' '.join(diagnosis)])
    medication = ''.join([' '.join(medication)])
    content = {
        'Progress Note' : progress,
        'Diagnosis' : diagnosis,
        'Lab Order' : lab,
        'Medication' : medication,
      
    }
    
    df = df.append(content, ignore_index=True)
    df.to_excel('steven2.xlsx')

    data1 = pd.read_excel('steven2.xlsx')
    features = ['Progress Note']
    df = data1[features]
    name = re.compile('\[.*\]')
   
    
    df['Progress Note'] = df['Progress Note'].apply(lambda x: x.replace(''.join(name.findall(x)), '') if type(x) is str else type(x))

    df['Progress Note'] = df['Progress Note'].astype(str)
    df['Progress Note'].apply(lambda x: ' '.join(x.lower() for x in x.split()))

    features  = df.loc[:,'Progress Note'].values
    
    nltk.download('stopwords')
    #df.text = df.text.apply(remove_stopwords)
    tfidf_vectorizer = TfidfVectorizer (max_features=600)
    tfidf_features = tfidf_vectorizer.fit_transform(features).toarray()
    print(tfidf_features.shape) 
   
    
    ynew = models.predict(tfidf_features)
    list1 = []
    print(ynew)

    for i in range(len(ynew)):
        if ynew[i] == 0:
            list1.append("Not ADR")
        elif ynew[i] == 1:
            list1.append("ADR")
        else:
            print( "Error")  
       
    data1['adr']  = pd.DataFrame(list1)
    data1['adr'].astype(str)
    data_to_save = ['Progress Note', 'adr']
    # data = pd.concat([data[features], my_predict_df], ignore_index=True) 

    data1[data_to_save].to_excel('file_name.xlsx')
    file_name = 'file_name.xlsx'  
    book = xlrd.open_workbook(file_name)
    sheet =book.sheet_by_index(0)

    cur = mysql.get_db().cursor()

    query = """INSERT INTO information(text,adr) VALUES (%s,%s)"""
    
    for r in range(1, sheet.nrows):
        text= sheet.cell(r,1).value
        adr= sheet.cell(r,2).value

        values = (text,adr)

        cur.execute(query,values)
    
    cur.close()

    mysql.get_db().commit()
        
    return redirect(url_for('downld'))

@app.route('/download', methods=['GET', 'POST'])
def downld():

   sql_select_Query = "select text,adr from information"
   cur = mysql.get_db().cursor()
   cur.execute(sql_select_Query)
   records = cur.fetchall()
   cur.close()

   return render_template('blank.html', records=records)



if __name__ == '__main__':
    app.debug = True
    app.run(debug=True)

