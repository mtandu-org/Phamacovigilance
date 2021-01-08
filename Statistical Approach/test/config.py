from flask import Flask, render_template, request, redirect, url_for , jsonify
from flaskext.mysql import MySQL
import xlrd
from openpyxl import load_workbook, Workbook
from werkzeug.utils import secure_filename
from sklearn.externals import joblib
import os
import pandas as pd
import json
import re
import nltk
import warnings
import numpy as np

import seaborn as sns
from argparse import Namespace
import matplotlib.pyplot as plt
from nltk.corpus import stopwords
from sklearn.utils import shuffle
from sklearn.feature_extraction.text import TfidfVectorizer


model_name = "KNN-f1-0.878"
models= joblib.load("../models/{}.{}".format(model_name,"pkl"))


app = Flask(__name__)


app.config['MYSQL_DATABASE_HOST'] = 'localhost'
app.config['MYSQL_DATABASE_USER'] = 'root'
app.config['MYSQL_DATABASE_PASSWORD'] = ''
app.config['MYSQL_DATABASE_DB'] = 'ADR_DEV'

mysql = MySQL(app)

@app.route('/')
def index():
    return render_template('blank.html')


@app.route('/upload', methods=['GET', 'POST'])
def up1():
    file = request.files['inputfile']
    file_name = 'temp.'+file.filename.split(".")[1]
    file.save(secure_filename(file_name))


    data1 = pd.read_excel(file_name)
    features = ['text']
    df = data1[features]
    name = re.compile('\[.*\]')
   
    df.text = df.text.str.strip()
    df['text'] = df['text'].apply(lambda x: x.replace(''.join(name.findall(x)), '') if type(x) is str else type(x))

    df['text'] = df['text'].astype(str)
    df['text'].apply(lambda x: ' '.join(x.lower() for x in x.split()))

    features  = df.loc[:,"text"].values
    
    nltk.download('stopwords')
    #df.text = df.text.apply(remove_stopwords)
    tfidf_vectorizer = TfidfVectorizer (max_features=600)
    tfidf_features = tfidf_vectorizer.fit_transform(features).toarray()
    print(tfidf_features.shape) 
   
    
    ynew = models.predict(tfidf_features)
    list1 = []
    print(ynew)

    for i in range(len(ynew)):
        if ynew[i] == 0:
            list1.append("Not ADR")
        elif ynew[i] == 1:
            list1.append("ADR")
        else:
            print( "Error")  
       
    data1['adr']  = pd.DataFrame(list1)
    data1['adr'].astype(str)
    data_to_save = ['text', 'adr']
    # data = pd.concat([data[features], my_predict_df], ignore_index=True) 

    data1[data_to_save].to_excel('file_name.xlsx')
    file_name = 'file_name.xlsx'  
    book = xlrd.open_workbook(file_name)
    sheet =book.sheet_by_index(0)

    cur = mysql.get_db().cursor()

    query = """INSERT INTO information(text,adr) VALUES (%s,%s)"""
    
    for r in range(1, sheet.nrows):
        text= sheet.cell(r,1).value
        adr= sheet.cell(r,2).value

        values = (text,adr)

        cur.execute(query,values)
    
    cur.close()

    mysql.get_db().commit()
        
    return redirect(url_for('downld'))

@app.route('/download', methods=['GET', 'POST'])
def downld():

   sql_select_Query = "select text,adr from information"
   cur = mysql.get_db().cursor()
   cur.execute(sql_select_Query)
   records = cur.fetchall()
   cur.close()

   return render_template('blank.html', records=records)



if __name__ == '__main__':
    app.debug = True
    app.run(debug=True)

cd 